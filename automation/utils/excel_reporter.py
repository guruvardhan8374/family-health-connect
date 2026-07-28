import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import REPORTS_DIR
from automation.utils.logger import logger

def create_styled_cell(cell, value, font_color="000000", bg_color=None, bold=False, align="left"):
    cell.value = value
    cell.font = Font(name="Segoe UI", size=11, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    cell.border = thin_border

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_excel_reports(test_results, metrics):
    logger.info("Generating Enterprise Excel Reports...")
    excel_dir = REPORTS_DIR / "Excel"
    excel_dir.mkdir(parents=True, exist_ok=True)

    # Master Report
    wb_master = openpyxl.Workbook()
    wb_master.remove(wb_master.active) # Remove default sheet

    headers = ["Test ID", "Module", "Test Name", "Priority", "Preconditions", "Test Steps", "Expected Result", "Actual Result", "Status", "Execution Time (s)"]

    # 1. Executed Test Cases
    ws_all = wb_master.create_sheet(title="Executed Test Cases")
    ws_all.append(headers)
    for col_num in range(1, len(headers) + 1):
        create_styled_cell(ws_all.cell(row=1, column=col_num), headers[col_num - 1], font_color="FFFFFF", bg_color="1E3A8A", bold=True, align="center")

    row_idx = 2
    for item in test_results:
        ws_all.append([
            item["id"], item["module"], item["name"], item["priority"],
            item.get("preconditions", ""), item.get("steps", ""),
            item.get("expected", ""), item.get("actual", ""),
            item["status"], round(item.get("duration", 0.1), 3)
        ])
        
        # Style Status Cell
        status_cell = ws_all.cell(row=row_idx, column=9)
        if item["status"] == "PASS":
            create_styled_cell(status_cell, "PASS", font_color="065F46", bg_color="D1FAE5", bold=True, align="center")
        elif item["status"] == "FAIL":
            create_styled_cell(status_cell, "FAIL", font_color="991B1B", bg_color="FEE2E2", bold=True, align="center")
        else:
            create_styled_cell(status_cell, "SKIPPED", font_color="92400E", bg_color="FEF3C7", bold=True, align="center")
        row_idx += 1

    auto_fit_columns(ws_all)

    # 2. Passed Tests Sheet
    ws_pass = wb_master.create_sheet(title="Passed Tests")
    ws_pass.append(headers)
    for col_num in range(1, len(headers) + 1):
        create_styled_cell(ws_pass.cell(row=1, column=col_num), headers[col_num - 1], font_color="FFFFFF", bg_color="065F46", bold=True, align="center")

    for item in [t for t in test_results if t["status"] == "PASS"]:
        ws_pass.append([
            item["id"], item["module"], item["name"], item["priority"],
            item.get("preconditions", ""), item.get("steps", ""),
            item.get("expected", ""), item.get("actual", ""),
            "PASS", round(item.get("duration", 0.1), 3)
        ])
    auto_fit_columns(ws_pass)

    # 3. Failed Tests Sheet
    ws_fail = wb_master.create_sheet(title="Failed Tests")
    ws_fail.append(headers)
    for col_num in range(1, len(headers) + 1):
        create_styled_cell(ws_fail.cell(row=1, column=col_num), headers[col_num - 1], font_color="FFFFFF", bg_color="991B1B", bold=True, align="center")

    for item in [t for t in test_results if t["status"] == "FAIL"]:
        ws_fail.append([
            item["id"], item["module"], item["name"], item["priority"],
            item.get("preconditions", ""), item.get("steps", ""),
            item.get("expected", ""), item.get("actual", ""),
            "FAIL", round(item.get("duration", 0.1), 3)
        ])
    auto_fit_columns(ws_fail)

    # 4. Skipped Tests Sheet
    ws_skip = wb_master.create_sheet(title="Skipped Tests")
    ws_skip.append(headers)
    for col_num in range(1, len(headers) + 1):
        create_styled_cell(ws_skip.cell(row=1, column=col_num), headers[col_num - 1], font_color="FFFFFF", bg_color="92400E", bold=True, align="center")

    for item in [t for t in test_results if t["status"] == "SKIPPED"]:
        ws_skip.append([
            item["id"], item["module"], item["name"], item["priority"],
            item.get("preconditions", ""), item.get("steps", ""),
            item.get("expected", ""), item.get("actual", ""),
            "SKIPPED", round(item.get("duration", 0.1), 3)
        ])
    auto_fit_columns(ws_skip)

    # 5. Execution Metrics Sheet
    ws_metrics = wb_master.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric Parameter", "Value"])
    create_styled_cell(ws_metrics.cell(row=1, column=1), "Metric Parameter", font_color="FFFFFF", bg_color="1E3A8A", bold=True)
    create_styled_cell(ws_metrics.cell(row=1, column=2), "Value", font_color="FFFFFF", bg_color="1E3A8A", bold=True)

    metric_rows = [
        ("Total Test Cases", metrics["total"]),
        ("Passed Tests", metrics["passed"]),
        ("Failed Tests", metrics["failed"]),
        ("Skipped Tests", metrics["skipped"]),
        ("Pass Percentage", f"{metrics['pass_rate']:.2f}%"),
        ("Execution Duration", f"{metrics['duration_sec']:.2f} seconds"),
        ("Deployment URL", metrics["base_url"]),
        ("Threshold Evaluation", metrics["threshold_status"])
    ]

    for r_idx, (k, v) in enumerate(metric_rows, start=2):
        ws_metrics.cell(row=r_idx, column=1, value=k)
        ws_metrics.cell(row=r_idx, column=2, value=str(v))

    auto_fit_columns(ws_metrics)

    # Save Master Report
    master_path = excel_dir / "Automation_Test_Report.xlsx"
    wb_master.save(master_path)

    # Save Passed_Test_Cases.xlsx
    wb_p = openpyxl.Workbook()
    ws_p = wb_p.active
    ws_p.title = "Passed Tests"
    ws_p.append(headers)
    for item in [t for t in test_results if t["status"] == "PASS"]:
        ws_p.append([item["id"], item["module"], item["name"], item["priority"], item.get("preconditions", ""), item.get("steps", ""), item.get("expected", ""), item.get("actual", ""), "PASS", round(item.get("duration", 0.1), 3)])
    wb_p.save(excel_dir / "Passed_Test_Cases.xlsx")

    # Save Failed_Test_Cases.xlsx
    wb_f = openpyxl.Workbook()
    ws_f = wb_f.active
    ws_f.title = "Failed Tests"
    ws_f.append(headers)
    for item in [t for t in test_results if t["status"] == "FAIL"]:
        ws_f.append([item["id"], item["module"], item["name"], item["priority"], item.get("preconditions", ""), item.get("steps", ""), item.get("expected", ""), item.get("actual", ""), "FAIL", round(item.get("duration", 0.1), 3)])
    wb_f.save(excel_dir / "Failed_Test_Cases.xlsx")

    # Save Summary_Report.xlsx
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.append(["Metric", "Value"])
    for k, v in metric_rows:
        ws_s.append([k, str(v)])
    wb_s.save(excel_dir / "Summary_Report.xlsx")

    logger.info(f"Excel Reports generated successfully in: {excel_dir}")
